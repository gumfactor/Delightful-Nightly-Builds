"""Export the quality report as a multi-sheet Excel workbook (.xlsx).

Requires openpyxl: pip install openpyxl
"""

import io
from datetime import datetime, timezone


# ── Colour palette (openpyxl ARGB hex, no leading #) ─────────────────────────
_BG_HEADER   = "FF1A1D27"
_BG_SURFACE  = "FF0F1117"
_FG_HEADER   = "FFE2E8F0"
_FG_MUTED    = "FF6B7280"
_FG_GREEN    = "FF22C55E"
_FG_AMBER    = "FFF59E0B"
_FG_RED      = "FFEF4444"
_FG_WHITE    = "FFE2E8F0"
_BG_GREEN    = "FF052E16"
_BG_AMBER    = "FF431407"
_BG_RED      = "FF450A0A"
_BG_GRAY     = "FF1A1D27"


def _styles():
    """Lazily import openpyxl styling classes."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        return PatternFill, Font, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")


def _make_wb():
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        return wb
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")


def _h(ws, row: int, col: int, value, bold: bool = True):
    """Write a header cell."""
    PatternFill, Font, Alignment = _styles()
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=_FG_HEADER, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=_BG_HEADER)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    return cell


def _d(ws, row: int, col: int, value, fg: str = _FG_WHITE, bg: str = _BG_SURFACE):
    """Write a data cell."""
    PatternFill, Font, Alignment = _styles()
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(color=fg, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return cell


def _auto_width(ws, min_width: int = 10, max_width: int = 60):
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        width = min_width
        for cell in col_cells:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def _rate_fg(rate: float) -> str:
    if rate < 0.05:
        return _FG_GREEN
    if rate < 0.20:
        return _FG_AMBER
    return _FG_RED


def _score_fg(score: float) -> str:
    if score < 0.2:
        return _FG_GREEN
    if score < 0.4:
        return _FG_AMBER
    return _FG_RED


# ── Sheet builders ────────────────────────────────────────────────────────────

def _add_overview(wb, quality, source_name: str):
    ws = wb.create_sheet("Overview")
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    thr = quality.thresholds

    _h(ws, 1, 1, "QUALTRICS SURVEY DATA QUALITY REPORT")
    _d(ws, 2, 1, f"Source: {source_name}",  fg=_FG_MUTED)
    _d(ws, 3, 1, f"Generated: {now}", fg=_FG_MUTED)

    headers = ["Metric", "Value", "Notes"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 5, ci, h)

    rows = [
        ("Total respondents",  quality.respondent_count, ""),
        ("Completed (100%)",   f"{quality.completed_count} ({quality.completion_rate:.1%})",
         f"< 90% is a concern"),
        (f"Fast responses (< {thr.fast_response_seconds}s)",
         len(quality.fast_response_ids), "Below timing threshold"),
        ("Straight-liners",    len(quality.straight_liner_ids), "Same answer to all scale items"),
        ("Duplicate IPs",      len(quality.duplicate_ips), "Potential duplicate submissions"),
        (f"High-missing respondents (> {thr.missing_respondent_flag:.0%})",
         len(quality.high_missing_respondents), "Missing > threshold of question items"),
        ("Columns with any missing", sum(1 for r in quality.per_column_missing.values() if r > 0), ""),
        ("Scales detected",    len(quality.detected_scales), ""),
        ("Outlier columns (Z or IQR)",
         len(quality.outlier_zscore) + len(quality.outlier_iqr), "Unique columns with outliers"),
    ]
    for ri, (label, val, note) in enumerate(rows, 6):
        _d(ws, ri, 1, label)
        _d(ws, ri, 2, val)
        _d(ws, ri, 3, note, fg=_FG_MUTED)

    _auto_width(ws)
    ws.freeze_panes = "A6"


def _add_respondents(wb, quality, survey, careless_index):
    ws = wb.create_sheet("Respondents")
    from src.quality import _get_response_id, _parse_duration

    fast_set = set(quality.fast_response_ids)
    straight_set = set(quality.straight_liner_ids)
    hm_set = set(quality.high_missing_respondents)
    dup_ips = set(quality.duplicate_ips)

    headers = ["ResponseId", "Progress", "Duration(s)", "Fast?", "Straight-liner?",
               "High Missing?", "Duplicate IP?", "Outlier Cols", "Careless Score", "Flags"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)

    for ri, row in enumerate(survey.rows, 2):
        rid = _get_response_id(row)
        progress = row.get("Progress") or ""
        duration = _parse_duration(row)
        is_fast = rid in fast_set
        is_straight = rid in straight_set
        is_hm = rid in hm_set
        ip = row.get("IPAddress") or row.get("ipAddress") or ""
        is_dup = ip in dup_ips
        n_outlier = quality.respondent_outlier_counts.get(rid, 0)

        careless = careless_index.get(rid) if careless_index else None
        c_score = careless["score"] if careless else 0.0
        c_flags = "|".join(careless["flags"]) if careless else ""

        _d(ws, ri, 1, rid)
        _d(ws, ri, 2, progress)
        _d(ws, ri, 3, duration)
        _d(ws, ri, 4, "YES" if is_fast else "", fg=_FG_RED if is_fast else _FG_MUTED)
        _d(ws, ri, 5, "YES" if is_straight else "", fg=_FG_RED if is_straight else _FG_MUTED)
        _d(ws, ri, 6, "YES" if is_hm else "", fg=_FG_AMBER if is_hm else _FG_MUTED)
        _d(ws, ri, 7, "YES" if is_dup else "", fg=_FG_AMBER if is_dup else _FG_MUTED)
        _d(ws, ri, 8, n_outlier if n_outlier else "", fg=_FG_AMBER if n_outlier else _FG_MUTED)
        _d(ws, ri, 9, round(c_score, 3) if c_score else "", fg=_score_fg(c_score))
        _d(ws, ri, 10, c_flags, fg=_FG_MUTED)

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _add_missing(wb, quality):
    ws = wb.create_sheet("Missing Data")

    _h(ws, 1, 1, "Column")
    _h(ws, 1, 2, "Missing Rate")
    _h(ws, 1, 3, "Severity")

    items = sorted(quality.per_column_missing.items(), key=lambda x: -x[1])
    for ri, (col, rate) in enumerate(items, 2):
        severity = "CRITICAL" if rate >= 0.20 else ("WARN" if rate >= 0.05 else "OK")
        fg = _FG_RED if rate >= 0.20 else (_FG_AMBER if rate >= 0.05 else _FG_GREEN)
        _d(ws, ri, 1, col)
        _d(ws, ri, 2, f"{rate:.1%}", fg=fg)
        _d(ws, ri, 3, severity, fg=fg)

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _add_reliability(wb, quality):
    ws = wb.create_sheet("Scale Reliability")

    _h(ws, 1, 1, "Scale")
    _h(ws, 1, 2, "Item")
    _h(ws, 1, 3, "Cronbach α")
    _h(ws, 1, 4, "α Quality")
    _h(ws, 1, 5, "Item-Total r")
    _h(ws, 1, 6, "r Flag")

    ri = 2
    for scale_name in sorted(quality.cronbach_results.keys()):
        alpha = quality.cronbach_results[scale_name]
        itc = quality.item_total_correlations.get(scale_name, {})
        items = quality.detected_scales.get(scale_name, [])
        label = _alpha_label(alpha) if alpha is not None else "n/a"
        alpha_str = f"{alpha:.3f}" if alpha is not None else "n/a"
        alpha_fg = _FG_GREEN if (alpha and alpha >= 0.70) else _FG_AMBER if alpha else _FG_MUTED

        for item in items:
            r = itc.get(item)
            low_r = r is not None and r < quality.thresholds.low_item_total_r
            r_fg = _FG_RED if low_r else (_FG_GREEN if r is not None and r >= 0.40 else _FG_WHITE)

            _d(ws, ri, 1, scale_name)
            _d(ws, ri, 2, item)
            _d(ws, ri, 3, alpha_str, fg=alpha_fg)
            _d(ws, ri, 4, label, fg=alpha_fg)
            _d(ws, ri, 5, f"{r:.3f}" if r is not None else "n/a", fg=r_fg)
            _d(ws, ri, 6, "LOW r !" if low_r else "", fg=_FG_RED if low_r else _FG_MUTED)
            ri += 1

        if items:
            ri += 1  # blank row between scales

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _add_normality(wb, quality):
    ws = wb.create_sheet("Normality")
    headers = ["Column", "N", "Mean", "SD", "Skewness", "Kurtosis", "K²", "p-value", "Normal?"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)

    ri = 2
    for col in sorted(quality.column_stats.keys()):
        s = quality.column_stats[col]
        norm = s.get("normality")
        if norm is None:
            continue
        is_normal = norm["is_normal"]
        p = norm["p_value"]
        p_fg = _FG_GREEN if is_normal else _FG_RED

        _d(ws, ri, 1, col)
        _d(ws, ri, 2, s.get("n"))
        _d(ws, ri, 3, s.get("mean"))
        _d(ws, ri, 4, s.get("std"))
        _d(ws, ri, 5, s.get("skewness"))
        _d(ws, ri, 6, s.get("kurtosis"))
        _d(ws, ri, 7, norm["statistic"])
        _d(ws, ri, 8, p, fg=p_fg)
        _d(ws, ri, 9, "Yes" if is_normal else "No *", fg=p_fg)
        ri += 1

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _add_outliers(wb, quality):
    ws = wb.create_sheet("Outliers")

    _h(ws, 1, 1, "Column")
    _h(ws, 1, 2, "Method")
    _h(ws, 1, 3, "ResponseId")
    _h(ws, 1, 4, "Value / Z-score")

    ri = 2
    for col, outliers in sorted(quality.outlier_zscore.items()):
        for rid, z in sorted(outliers.items(), key=lambda x: -abs(x[1])):
            _d(ws, ri, 1, col)
            _d(ws, ri, 2, "Z-score", fg=_FG_AMBER)
            _d(ws, ri, 3, rid)
            _d(ws, ri, 4, z, fg=_FG_RED if abs(z) > 4 else _FG_AMBER)
            ri += 1

    for col, outliers in sorted(quality.outlier_iqr.items()):
        for rid, val in sorted(outliers.items()):
            _d(ws, ri, 1, col)
            _d(ws, ri, 2, "IQR", fg=_FG_AMBER)
            _d(ws, ri, 3, rid)
            _d(ws, ri, 4, val, fg=_FG_AMBER)
            ri += 1

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _add_attention(wb, attention_results: dict):
    ws = wb.create_sheet("Attention Checks")
    headers = ["Column", "Expected Answer", "N Checked", "Pass Rate", "N Failed", "Failed IDs"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 1, ci, h)

    for ri, (col, res) in enumerate(sorted(attention_results.items()), 2):
        expected = res.get("expected") or "(unknown)"
        n_checked = res.get("n_checked", 0)
        pass_rate = res.get("pass_rate")
        failed_ids = res.get("failed_ids", [])
        n_failed = len(failed_ids)

        pr_fg = _FG_MUTED if pass_rate is None else (_FG_GREEN if pass_rate >= 0.9 else _FG_AMBER if pass_rate >= 0.7 else _FG_RED)
        pr_str = f"{pass_rate:.1%}" if pass_rate is not None else "n/a"

        _d(ws, ri, 1, col)
        _d(ws, ri, 2, expected, fg=_FG_MUTED)
        _d(ws, ri, 3, n_checked)
        _d(ws, ri, 4, pr_str, fg=pr_fg)
        _d(ws, ri, 5, n_failed, fg=_FG_RED if n_failed else _FG_GREEN)
        _d(ws, ri, 6, ", ".join(failed_ids[:20]) + ("…" if len(failed_ids) > 20 else ""),
           fg=_FG_MUTED)

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _add_careless(wb, careless_index: dict, threshold: float = 0.4):
    from src.careless import careless_summary
    ws = wb.create_sheet("Careless Index")

    summary = careless_summary(careless_index, threshold)

    _h(ws, 1, 1, "Summary")
    _d(ws, 2, 1, f"Flagged respondents (score ≥ {threshold})")
    _d(ws, 2, 2, summary["n_flagged"], fg=_FG_RED if summary["n_flagged"] else _FG_GREEN)
    _d(ws, 3, 1, "Mean careless score")
    _d(ws, 3, 2, summary["mean_score"])

    _h(ws, 5, 1, "Score distribution")
    _h(ws, 5, 2, "Count")
    for ri, (band, count) in enumerate(summary["score_distribution"].items(), 6):
        _d(ws, ri, 1, band, fg=_score_fg(float(band.split("–")[0])))
        _d(ws, ri, 2, count)

    headers = ["ResponseId", "Score", "Fast?", "Straight-liner?", "High-missing?",
               "Outlier breadth", "Attn fail rate", "Flags"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 13, ci, h)

    rows_data = sorted(careless_index.items(), key=lambda x: -x[1]["score"])
    for ri, (rid, data) in enumerate(rows_data, 14):
        score = data["score"]
        comp = data["components"]
        _d(ws, ri, 1, rid)
        _d(ws, ri, 2, round(score, 3), fg=_score_fg(score))
        _d(ws, ri, 3, round(comp.get("fast_response", 0), 2))
        _d(ws, ri, 4, round(comp.get("straight_liner", 0), 2))
        _d(ws, ri, 5, round(comp.get("high_missing", 0), 2))
        _d(ws, ri, 6, round(comp.get("outlier_breadth", 0), 3))
        _d(ws, ri, 7, round(comp.get("attention_fail_rate", 0), 3)
           if "attention_fail_rate" in comp else "n/a")
        _d(ws, ri, 8, "|".join(data["flags"]), fg=_FG_MUTED)

    _auto_width(ws)
    ws.freeze_panes = "A14"


def _add_group_tests(wb, quality):
    ws = wb.create_sheet("Group Tests")

    _h(ws, 1, 1, f"Condition column: {quality.condition_column}")
    headers = ["Scale", "Group", "N", "Mean", "SD", "Median", "Test type", "Statistic", "p-value", "Significant?"]
    for ci, h in enumerate(headers, 1):
        _h(ws, 2, ci, h)

    ri = 3
    for scale, result in sorted(quality.condition_tests.items()):
        test = result.get("test") or {}
        t_type = test.get("type", "")
        stat_key = "H" if "Kruskal" in t_type else "U"
        stat_val = test.get(stat_key, "")
        p = test.get("p_value")
        sig = "YES *" if p is not None and p < 0.05 else "no"
        p_fg = _FG_RED if (p is not None and p < 0.05) else _FG_GREEN

        for group, gs in sorted(result.get("group_stats", {}).items()):
            _d(ws, ri, 1, scale)
            _d(ws, ri, 2, str(group))
            _d(ws, ri, 3, gs.get("n"))
            _d(ws, ri, 4, gs.get("mean"))
            _d(ws, ri, 5, gs.get("std"))
            _d(ws, ri, 6, gs.get("median"))
            _d(ws, ri, 7, t_type, fg=_FG_MUTED)
            _d(ws, ri, 8, stat_val, fg=_FG_MUTED)
            _d(ws, ri, 9, p, fg=p_fg)
            _d(ws, ri, 10, sig, fg=p_fg)
            ri += 1
        ri += 1

    _auto_width(ws)
    ws.freeze_panes = "A3"


def _add_correlation_sheet(wb, scale_name: str, matrix: dict):
    sheet_name = f"Corr {scale_name}"[:31]  # Excel sheet name limit
    ws = wb.create_sheet(sheet_name)
    cols = list(matrix.keys())

    # Header row
    _h(ws, 1, 1, scale_name)
    for ci, col in enumerate(cols, 2):
        _h(ws, 1, ci, col[:20])

    # Data rows
    for ri, row_col in enumerate(cols, 2):
        _h(ws, ri, 1, row_col[:20])
        for ci, col_col in enumerate(cols, 2):
            val = matrix[row_col].get(col_col)
            if row_col == col_col:
                _d(ws, ri, ci, "1.000", fg=_FG_MUTED)
            elif val is None:
                _d(ws, ri, ci, "n/a", fg=_FG_MUTED)
            else:
                fg = _corr_fg(val)
                _d(ws, ri, ci, round(val, 3), fg=fg)

    _auto_width(ws, min_width=8, max_width=20)
    ws.freeze_panes = "B2"


def _corr_fg(r: float) -> str:
    if r >= 0.70:
        return _FG_GREEN
    if r >= 0.30:
        return "FF86EFAC"  # light green
    if r <= -0.70:
        return _FG_RED
    if r <= -0.30:
        return "FFFCA5A5"  # light red
    return _FG_WHITE


def _alpha_label(alpha: float) -> str:
    if alpha >= 0.90:
        return "excellent"
    if alpha >= 0.80:
        return "good"
    if alpha >= 0.70:
        return "acceptable"
    if alpha >= 0.60:
        return "questionable"
    if alpha >= 0.50:
        return "poor"
    return "unacceptable"


# ── Main entry point ──────────────────────────────────────────────────────────

def export_excel(
    quality,
    survey,
    source_name: str = "survey",
    attention_results: dict = None,
    careless_index: dict = None,
) -> bytes:
    """
    Generate a multi-sheet Excel workbook and return its bytes.

    Sheets created:
      Overview         — summary stats
      Respondents      — per-respondent flags + careless score
      Missing Data     — per-column missing rates
      Scale Reliability — Cronbach α + item-total correlations
      Normality        — skewness, kurtosis, K², p-value per column
      Outliers         — Z-score and IQR outlier details (if any)
      Attention Checks — pass/fail per check column (if any)
      Careless Index   — composite score per respondent (if provided)
      Group Tests      — between-group test results (if conditions detected)
      Corr {scale}     — correlation matrix per scale

    Raises ImportError if openpyxl is not installed.
    """
    wb = _make_wb()

    _add_overview(wb, quality, source_name)
    _add_respondents(wb, quality, survey, careless_index)
    _add_missing(wb, quality)
    _add_reliability(wb, quality)
    _add_normality(wb, quality)

    if quality.outlier_zscore or quality.outlier_iqr:
        _add_outliers(wb, quality)

    if attention_results:
        _add_attention(wb, attention_results)

    if careless_index:
        _add_careless(wb, careless_index)

    if quality.condition_tests:
        _add_group_tests(wb, quality)

    for scale_name, matrix in quality.correlation_matrices.items():
        _add_correlation_sheet(wb, scale_name, matrix)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
